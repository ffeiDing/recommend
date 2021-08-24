# 先清洗数据，仅留下jsonb的action为checkedout的条目

import openpyxl
import json

workbook = openpyxl.load_workbook("data/loan.xlsx")
shenames = workbook.sheetnames
# print(shenames)

worksheet = workbook["result 1"]
# print(worksheet) 

rows = worksheet.max_row
columns = worksheet.max_column
print(rows, columns) # 346000 4

clean_rows = []
first_row = True
for row in worksheet.rows:
    if first_row:
        first_row = False
        continue
    clean_row = []
    jsonb = json.loads(row[1].value)
    # print(jsonb["action"])
    if jsonb["action"] != "checkedout":
        continue
    for cell in row:
        clean_row.append(cell.value)
    # print(clean_row)
    if len(clean_row) != 0:
        clean_rows.append(clean_row)


clean_worksheet = workbook.create_sheet() 
clean_worksheet.title = "checkedout"

for i in range(len(clean_rows)):
    for j in range(columns):
        clean_worksheet.cell(i+1, j+1, clean_rows[i][j])

clean_rows = []
first_row = True
for row in worksheet.rows:
    if first_row:
        first_row = False
        continue
    clean_row = []
    jsonb = json.loads(row[1].value)
    # print(jsonb["action"])
    if jsonb["action"] != "checkedin":
        continue
    for cell in row:
        clean_row.append(cell.value)
    # print(clean_row)
    if len(clean_row) != 0:
        clean_rows.append(clean_row)


clean_worksheet = workbook.create_sheet() 
clean_worksheet.title = "checkedin"

for i in range(len(clean_rows)):
    for j in range(columns):
        clean_worksheet.cell(i+1, j+1, clean_rows[i][j])


workbook.save(filename='data/loan.xlsx')