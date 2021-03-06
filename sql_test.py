import pymysql
import json
import requests
import openpyxl

server = "127.0.0.1"    # 数据库服务器名称或IP
user = "root"   #  用户名
password = "284284dfl" # 密码
database =  "loan_new" # 数据库名称
database_ckey =  "ckey2title" # 数据库名称

def get_book_main_callnumber_by_ckey(book_ckey):
    callnumber = ""
    return callnumber[0]

# 根据书籍ID获取书名
def get_book_info(book_id):
    if "/" in book_id:
        return "No Title!"
    url = "http://162.105.134.104:9130/inventory/items?query=barcode="+book_id
    headers = {"X-Okapi-Tenant":"pkulib", "X-Okapi-Token":"eyJhbGciOiJIUzI1NiJ9.eyJzdWIiOiJwa3VsaWIiLCJ1c2VyX2lkIjoiNzA2OWE3YjAtMzZlNy00ZjlkLWJjNDItMmQwNjQwYWIyNjhkIiwiaWF0IjoxNjI5Mjc3NzA3LCJ0ZW5hbnQiOiJwa3VsaWIifQ.utKpvgovU943UX_KnOIGzObzcwh07cijaHT8cjjJREA"}
    response = requests.get(url, headers=headers)
    # print(response.headers)
    result = response.text
    # print(book_id)
    book_info = json.loads(result)["items"]
    book_title = book_info[0]["title"]
    return book_title

# 根据书籍ckey获取书名
def get_book_info_by_ckey(book_ckey, cur):
    sql = "SELECT ckey, title FROM bibli WHERE ckey='"+book_ckey+"'"
    cur.execute(sql)
    book_info = cur.fetchall()
    if len(book_info) == 0:
        return "None Title"
    return book_info[0][1]

# 根据id list获取书籍title list
def get_recommend_book(book_id_list):
    book_title_list = []
    for book_id in book_id_list:
        book_id = book_id[0]
        book_title = get_book_info(book_id)
        book_title_list.append(book_title)
        # print(book_title)
    return book_title_list

# 根据ckey list获取书籍title list
def get_book_title_list_by_ckey_list(book_ckey_list):
    db = pymysql.connect(host=server, user=user, passwd=password, db=database_ckey)
    cur_ckey = db.cursor()
    book_title_list = []
    for book_ckey in book_ckey_list:
        book_ckey = book_ckey[0]
        # print(book_ckey)
        book_title = get_book_info_by_ckey(book_ckey, cur_ckey)
        if len(book_title.strip()) == 0:
            continue
        book_title_list.append(book_title)
        # print(book_title)
    return book_title_list


# 获取用户历史借阅书籍列表
def get_user_history(user_id):
    book_title_list = []
    db = pymysql.connect(host=server, user=user, passwd=password, db=database)
    cur = db.cursor()
    sql = "SELECT altid, item_id, ckey FROM userlog WHERE altid="+user_id
    cur.execute(sql) #共7974414条记录
    for line in cur.fetchall():
        book_id = line[1]
        book_title = get_book_info(book_id)
        if book_title == "No Title!" or book_title == "馆际互借代借国家图书馆图书":
            continue
        book_title_list.append(book_title)
    return book_title_list


# 根据user_id获取用户历史借阅书籍列表
def get_history_book_title_list_by_user_id(user_id):
    book_title_list = []
    db = pymysql.connect(host=server, user=user, passwd=password, db=database)
    cur = db.cursor()
    sql = "SELECT altid, item_id, ckey FROM userlog WHERE altid="+user_id
    cur.execute(sql) #共7974414条记录
    db_ckey = pymysql.connect(host=server, user=user, passwd=password, db=database_ckey)
    cur_ckey = db_ckey.cursor()
    for line in cur.fetchall():
        book_ckey = line[2]
        book_title = get_book_info_by_ckey(book_ckey, cur_ckey)
        if book_title == "No Title!" or book_title == "馆际互借代借国家图书馆图书":
            continue
        book_title_list.append(book_title)
    return book_title_list

# 根据用户名获取用户ID
def get_user_id_by_name(name):
    db = pymysql.connect(host=server, user=user, passwd=password, db=database)
    cur = db.cursor()
    sql = "SELECT altid, name, item_id, ckey FROM userlog WHERE name = '" + name + "'"
    cur.execute(sql)
    return cur.fetchall()[0][0]
    

# 根据书籍条码获得ckey
def get_ckey_by_book_id(book_id):
    db = pymysql.connect(host=server, user=user, passwd=password, db=database)
    cur = db.cursor()
    sql = "SELECT altid, item_id, ckey FROM userlog WHERE item_id='"+book_id+"'"
    cur.execute(sql) #共7974414条记录
    book_info = cur.fetchall()
    if len(book_info) == 0:
        return ""
    return book_info[0][2]

# 根据书籍条码list获取ckey_list
def get_ckey_list_by_book_id_list(book_id_list):
    ckey_list = []
    for book_id in book_id_list:
        book_ckey = get_ckey_by_book_id(book_id)
        if len(book_ckey.strip()) == 0:
            continue
        ckey_list.append(book_ckey)
    return ckey_list

# 读取用户xls格式的借书记录，返回借阅过图书的id list
def get_user_history_book_id_list_from_xls(file_name):
    first_row = True
    book_id_list = []
    workbook = openpyxl.load_workbook(file_name)
    sheetnames = workbook.sheetnames
    sheet = workbook[sheetnames[0]]
    for row in sheet.rows:
        if first_row:
            first_row = False
            continue
        book_id = row[4].value
        book_id_list.append(book_id)
    book_id_list = book_id_list[: :-1]
    return book_id_list 

# 读取用户xls格式的借书记录，返回借阅过图书的title list
def get_user_history_book_title_list_from_xls(file_name):
    first_row = True
    book_title_list = []
    workbook = openpyxl.load_workbook(file_name)
    sheetnames = workbook.sheetnames
    sheet = workbook[sheetnames[0]]
    for row in sheet.rows:
        if first_row:
            first_row = False
            continue
        book_title = row[3].value
        book_title_list.append(book_title)
    book_title_list = book_title_list[: :-1]
    return book_title_list 

if __name__ == "__main__":
    # get_user_id_by_name("王昊贤")
    # print(get_recommend_book_by_ckey(["012000905826"]))
    print(get_history_book_title_list_by_user_id("1400012962"))
