# 先清洗数据，仅留下jsonb的action为checkedout的条目
import jieba
import openpyxl
import json
import pickle

def load_dic(name):
    with open(name, 'rb') as f:
        return pickle.load(f)   

workbook = openpyxl.load_workbook("data/daka.xlsx")
shenames = workbook.sheetnames
# print(shenames)

worksheet = workbook["result 1"]
# print(worksheet) 

rows = worksheet.max_row
columns = worksheet.max_column
# print(rows, columns) # 346000 4

daka = {}
first_row = True
for row in worksheet.rows:
    if first_row:
        first_row = False
        continue
    user_id = row[1].value
    daka.setdefault(user_id,{})
    level = row[2].value
    daka[user_id][level] = 1
    if level == 3:
        daka[user_id]["book_name"] = row[4].value


reader = load_dic("data/reader_info.pkl")

cnt = 0
g = {}
s = {}
y = {}
book_list = " "
med_book_list = " "
finish_all = {}
tmp = 0
daka_cnt = [0 for i in range(6)]
for key, user_info in daka.items():
    school = ""
    if key in reader.keys():
        reader_info = reader[key]
        gender = reader_info["gender"]
        if gender == '\n':
            print(key)
            continue
        cnt += 1
        school =  reader_info["school"]
        if "book_name" in user_info.keys():
            book_list += user_info["book_name"].split()[0]
        if school == "信息技术" and "book_name" in user_info.keys():
            med_book_list += user_info["book_name"].split()[0]
            tmp = tmp+1
        k = key[:2]
        y.setdefault(k, 0)
        y[k] += 1
        g.setdefault(gender, 0)
        g[gender] += 1
        s.setdefault(school, 0)
        s[school] += 1
        # print(reader[key])
    # else:
    #     print(key)
    meet_all = True
    for i in range(5):
        if i+1 in user_info.keys():
            daka_cnt[i] += 1
        else:
            meet_all = False
    if meet_all == True:
        daka_cnt[5] += 1
        finish_all.setdefault(school, 0)
        finish_all[school] += 1

print("参与人数：", len(daka))
print("全部通关人数：", daka_cnt[5])
for i in range(5):
    print("通过第" + str(i+1) + "关的人数：", daka_cnt[i])
print(cnt)
print(g, s)
for key, value in s.items():
    print(key, value)
print(y)
print(book_list)
# seg_list = jieba.cut(book_list, cut_all=False)
# res = ' '.join(seg_list).split()
# print(res)
# dic = {}
# for i in res:
#     dic.setdefault(i, 0)
#     dic[i] += 1
# for key, value in dic.items():
#     print(key, value)

# for key, value in finish_all.items():
#     print(key, value)


seg_list = jieba.cut(med_book_list, cut_all=False)
res = ' '.join(seg_list).split()
print(res)
dic = {}
for i in res:
    dic.setdefault(i, 0)
    dic[i] += 1
for key, value in dic.items():
    print(key, value)
print(tmp)

